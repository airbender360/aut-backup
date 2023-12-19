import openpyxl
from copy import copy
from openpyxl.worksheet.table import Table, TableStyleInfo

class Hoja:
    def __init__(self, ruta, nombre, datos):
        self.ruta = ruta + r'\bd.xlsx'
        self.nombre = nombre
        self.datos = datos
        self.libro = openpyxl.load_workbook(self.ruta)
        self.hojaPlantilla = openpyxl.load_workbook(ruta + r'\plantilla.xlsx').active
        if self.nombre in self.libro.sheetnames:
            print('Error: La hoja ' + self.nombre + " ya existe")
        else:
            self.hoja = self.libro.create_sheet(self.nombre)
        
    def crearRegistro(self):
        try:
            hoja = self.hoja
            for indice, registro in enumerate(self.datos):
                hoja['A' + str(indice+2)] = self.nombre
                hoja['B' + str(indice+2)] = registro['YouTubeName']
                hoja['C' + str(indice+2)] = registro['UploadDate']
                hoja['D' + str(indice+2)] = registro['Duration']
                hoja['E' + str(indice+2)] = registro['FileSize']
                hoja['F' + str(indice+2)] = registro['ImageHeight']
                hoja['G' + str(indice+2)] = registro['FileTypeExtension']
                hoja['H' + str(indice+2)] = 'Santiago Sánchez'
                hoja['I' + str(indice+2)] = registro['FileCreateDate']
                
            self.libro.save(self.ruta)
            self.hoja = hoja
        except Exception as e: 
            print(f'Error: {e}')
        finally:
            print('6. Documentación exitosa (Excel)')
        
    def formatearHoja(self):
        self.copiarCeldas(self.hojaPlantilla, self.hoja) 
        self.crearTabla(self.hoja)
    
    def copiarCeldas(self, plantillaHoja, nuevaHoja):
        for (row, col), plantillaCelda in plantillaHoja._cells.items():
            nuevaCelda = nuevaHoja.cell(column=col, row=row)
            nuevaCelda._value = plantillaCelda._value
            nuevaCelda.data_type = plantillaCelda.data_type
            
        for indice, value in plantillaHoja.column_dimensions.items():
            nuevaHoja.column_dimensions[indice].width = copy(plantillaHoja.column_dimensions[indice].width)
        
    def crearTabla(self, hojaNueva):
        rango = len(self.datos)
        rangoTabla = f"A1:L{rango+1}"
        tabla = Table(displayName=f"tbl{self.nombre.replace(" ", "")}", ref=rangoTabla)
        estilo = TableStyleInfo(
            name="TableStyleMedium3", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=True
        )
        tabla.tableStyleInfo = estilo
        hojaNueva.add_table(tabla)